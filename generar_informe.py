"""
Script para generar informe de funcionalidades del sistema ERP-PSI en Excel
"""
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from datetime import date

# ─────────────────────────────────────────────
# DATOS DEL INFORME
# ─────────────────────────────────────────────
SISTEMA = "ERP-PSI — Sistema de Gestión para Proveedor de Internet"
FECHA   = date.today().strftime("%d/%m/%Y")

funcionalidades = [
    # (N°, Módulo, Funcionalidad, Descripción, Estado, Observaciones)

    # AUTH & USUARIOS
    (1, "Autenticación y Usuarios", "Login / Logout",
     "Inicio y cierre de sesión con tokens JWT",
     "Terminado", ""),
    (2, "Autenticación y Usuarios", "Registro de usuarios internos",
     "Creación de cuentas para administradores, supervisores, instaladores y técnicos",
     "Terminado", ""),
    (3, "Autenticación y Usuarios", "Control de acceso por roles",
     "Permisos diferenciados según rol (admin, supervisor, instalador, técnico)",
     "Terminado", ""),
    (4, "Autenticación y Usuarios", "Cambio y recuperación de contraseña",
     "Flujo de cambio de contraseña; recuperación pendiente de integrar envío de correo",
     "En proceso", "Falta completar el envío de email en recuperación"),
    (5, "Autenticación y Usuarios", "Gestión de perfil de usuario",
     "Consulta y edición de datos del usuario autenticado",
     "Terminado", ""),

    # CLIENTES
    (6, "Gestión de Clientes", "CRUD de clientes",
     "Crear, consultar, editar y eliminar clientes",
     "Terminado", ""),
    (7, "Gestión de Clientes", "Soporte multi-sede",
     "Un cliente con el mismo ID puede tener varias ubicaciones/sedes con servicios independientes",
     "Terminado", "Migración de BD completada"),
    (8, "Gestión de Clientes", "Verificación y alertas de duplicados",
     "Detección automática de clientes existentes al momento del registro",
     "Terminado", ""),
    (9, "Gestión de Clientes", "Filtros, búsqueda y exportación",
     "Búsqueda por nombre, documento, ciudad; exportación de listados",
     "Terminado", ""),
    (10, "Gestión de Clientes", "Gestión de clientes inactivos",
     "Archivo y consulta de clientes dados de baja",
     "Terminado", ""),
    (11, "Gestión de Clientes", "Estadísticas de clientes",
     "Indicadores de clientes activos, inactivos y crecimiento",
     "Terminado", ""),
    (12, "Gestión de Clientes", "Registro público web",
     "Formulario de registro sin autenticación para prospectos",
     "Terminado", ""),
    (13, "Gestión de Clientes", "Consulta pública de estado",
     "Portal para que el cliente consulte su estado de cuenta sin login",
     "Terminado", ""),

    # FACTURACIÓN
    (14, "Facturación e Invoices", "Ciclo completo de facturas",
     "Creación, consulta, edición y anulación de facturas",
     "Terminado", ""),
    (15, "Facturación e Invoices", "Numeración automática de facturas",
     "Procedimiento almacenado que genera número único de factura",
     "Terminado", ""),
    (16, "Facturación e Invoices", "Facturación automática mensual",
     "Cron job el día 1 de cada mes genera facturas de todos los servicios activos",
     "Terminado", ""),
    (17, "Facturación e Invoices", "Factura proporcional (primer mes)",
     "Cálculo proporcional de días desde la activación hasta fin de mes",
     "Terminado", ""),
    (18, "Facturación e Invoices", "Cálculo de IVA por estrato",
     "Exención de IVA para estratos 1-3 en servicio de internet según normativa colombiana",
     "Terminado", ""),
    (19, "Facturación e Invoices", "Intereses moratorios",
     "Cálculo automático de intereses sobre facturas vencidas",
     "Terminado", ""),
    (20, "Facturación e Invoices", "Facturación masiva",
     "Generación en lote de facturas para múltiples clientes simultáneamente",
     "Terminado", ""),
    (21, "Facturación e Invoices", "Generación de PDF de factura",
     "PDF con código de barras y firma digital",
     "Terminado", ""),
    (22, "Facturación e Invoices", "Seguimiento de pagos",
     "Registro de pagos, métodos de pago y conciliación",
     "Terminado", ""),
    (23, "Facturación e Invoices", "Integración formatos bancarios",
     "Soporte Asobancaria, Cajasocial, Comultrasan, Finecoop",
     "Terminado", ""),

    # CONTRATOS
    (24, "Contratos", "CRUD de contratos",
     "Creación, consulta y gestión del ciclo de vida del contrato",
     "Terminado", ""),
    (25, "Contratos", "Generación automática al registrar cliente",
     "Contrato generado automáticamente al completar el registro",
     "Terminado", ""),
    (26, "Contratos", "PDF de contrato con firma digital",
     "Generación del documento PDF con soporte de firma electrónica",
     "Terminado", ""),
    (27, "Contratos", "Plantilla MINTIC",
     "Modelo de contrato alineado a requisitos del regulador colombiano",
     "Terminado", ""),
    (28, "Contratos", "Numeración automática de contratos",
     "Procedimiento almacenado para número único de contrato",
     "Terminado", ""),

    # INSTALACIONES
    (29, "Gestión de Instalaciones", "Órdenes de instalación",
     "Creación, asignación y seguimiento de órdenes",
     "Terminado", ""),
    (30, "Gestión de Instalaciones", "Programación y calendario",
     "Calendario visual para planificar instalaciones",
     "Terminado", ""),
    (31, "Gestión de Instalaciones", "Asignación de técnicos",
     "Asignación de instaladores a cada orden",
     "Terminado", ""),
    (32, "Gestión de Instalaciones", "Carga de evidencias fotográficas",
     "Registro de fotos y documentación de la instalación",
     "Terminado", ""),
    (33, "Gestión de Instalaciones", "Mapa de instalaciones",
     "Visualización geográfica de instalaciones con Leaflet",
     "Terminado", ""),
    (34, "Gestión de Instalaciones", "Vista del instalador (app móvil-web)",
     "Interfaz simplificada para que el instalador vea y gestione sus trabajos",
     "Terminado", ""),

    # INVENTARIO
    (35, "Inventario de Equipos", "CRUD de equipos",
     "Registro, consulta y edición de equipos (routers, módems, cables, etc.)",
     "Terminado", ""),
    (36, "Inventario de Equipos", "Estados de equipo",
     "Control de estado: disponible, asignado, instalado, devuelto",
     "Terminado", ""),
    (37, "Inventario de Equipos", "Asignación a instaladores",
     "Entrega y devolución de equipos por instalador",
     "Terminado", ""),
    (38, "Inventario de Equipos", "Historial y auditoría",
     "Trazabilidad completa de movimientos de cada equipo",
     "Terminado", ""),
    (39, "Inventario de Equipos", "Equipos perdidos",
     "Registro y seguimiento de equipos reportados como perdidos",
     "Terminado", ""),
    (40, "Inventario de Equipos", "Reportes de disponibilidad",
     "Informes de stock disponible y equipos en campo",
     "Terminado", ""),

    # PLANES DE SERVICIO
    (41, "Planes de Servicio", "CRUD de planes",
     "Creación y gestión de planes de internet, TV y combinados",
     "Terminado", ""),
    (42, "Planes de Servicio", "Activación/desactivación de planes",
     "Control de vigencia de cada plan",
     "Terminado", ""),
    (43, "Planes de Servicio", "Conceptos de facturación",
     "Gestión de conceptos (instalación, servicio, mora, etc.) asociados a planes",
     "Terminado", ""),

    # CONFIGURACIÓN DEL SISTEMA
    (44, "Configuración del Sistema", "Datos de la empresa",
     "Información corporativa usada en documentos y reportes",
     "Terminado", ""),
    (45, "Configuración del Sistema", "Configuración geográfica",
     "Gestión de departamentos, ciudades, sectores y zonas",
     "Terminado", ""),
    (46, "Configuración del Sistema", "Configuración de bancos",
     "Datos bancarios y formatos de pago por entidad",
     "Terminado", ""),
    (47, "Configuración del Sistema", "Parámetros de facturación",
     "Tasas de mora, fechas de corte, IVA configurable",
     "Terminado", ""),
    (48, "Configuración del Sistema", "Plantillas de correo electrónico",
     "Editor de plantillas para notificaciones por email",
     "Terminado", ""),
    (49, "Configuración del Sistema", "Respaldo (backup) de base de datos",
     "Generación y descarga de backup desde la interfaz",
     "Terminado", ""),
    (50, "Configuración del Sistema", "Logs del sistema",
     "Registro de auditoría de acciones críticas",
     "Terminado", ""),

    # PQR
    (51, "PQR (Peticiones, Quejas y Reclamos)", "Creación y seguimiento de PQR",
     "Registro, asignación y resolución de solicitudes del cliente",
     "Terminado", ""),
    (52, "PQR (Peticiones, Quejas y Reclamos)", "Estados y escalamiento",
     "Flujo de estados con posibilidad de escalar al nivel superior",
     "Terminado", ""),
    (53, "PQR (Peticiones, Quejas y Reclamos)", "Reportes estadísticos de PQR",
     "Indicadores de tipos, tiempos de resolución y tendencias",
     "Terminado", ""),

    # INCIDENCIAS
    (54, "Gestión de Incidencias", "Registro de incidencias de servicio",
     "Apertura y categorización de fallos de red o servicio",
     "Terminado", ""),
    (55, "Gestión de Incidencias", "Prioridades y asignación",
     "Niveles de prioridad y asignación a técnico responsable",
     "Terminado", ""),
    (56, "Gestión de Incidencias", "Resolución y cierre",
     "Registro de solución aplicada y cierre de incidencia",
     "Terminado", ""),
    (57, "Gestión de Incidencias", "Estadísticas de incidencias",
     "Indicadores de disponibilidad y tiempo medio de resolución",
     "Terminado", ""),

    # NOTIFICACIONES
    (58, "Notificaciones", "Campana de notificaciones en tiempo real",
     "Icono con contador de no leídas; auto-refresco cada 30 s",
     "Terminado", ""),
    (59, "Notificaciones", "Generación automática de alertas",
     "Notificación al crear cliente, orden de instalación, PQR, etc.",
     "Terminado", ""),
    (60, "Notificaciones", "Notificaciones por rol",
     "Cada rol recibe solo las notificaciones pertinentes",
     "Terminado", ""),
    (61, "Notificaciones", "Notificaciones por correo electrónico",
     "Envío de emails automáticos con plantillas configurables",
     "Terminado", ""),

    # SOPORTE / CHATBOT
    (62, "Soporte al Cliente", "Página pública de soporte",
     "Acceso sin login para clientes con preguntas frecuentes y chatbot",
     "Terminado", ""),
    (63, "Soporte al Cliente", "Chatbot con IA (Google Gemini)",
     "Asistente conversacional con respaldo en API de Gemini",
     "Terminado", ""),
    (64, "Soporte al Cliente", "Creación automática de PQR desde chat",
     "Si el chatbot no resuelve, genera un ticket de soporte automáticamente",
     "Terminado", ""),

    # REPORTES REGULATORIOS
    (65, "Reportes Regulatorios (DIAN/MINTIC)", "Informe de suscriptores activos",
     "Reporte de abonados activos en el período",
     "Terminado", ""),
    (66, "Reportes Regulatorios (DIAN/MINTIC)", "Informe de planes de precios",
     "Líneas tarifarias y valores por plan",
     "Terminado", ""),
    (67, "Reportes Regulatorios (DIAN/MINTIC)", "Métricas QoS",
     "Indicadores de calidad del servicio (disponibilidad, latencia)",
     "Terminado", ""),
    (68, "Reportes Regulatorios (DIAN/MINTIC)", "Monitoreo de PQRS",
     "Reporte regulatorio de quejas y reclamaciones",
     "Terminado", ""),
    (69, "Reportes Regulatorios (DIAN/MINTIC)", "Facturación y ventas",
     "Reporte de facturas emitidas para declaración ante DIAN",
     "Terminado", ""),

    # DASHBOARD
    (70, "Dashboard y Analítica", "Panel principal",
     "Vista resumen con KPIs de clientes, facturas e instalaciones",
     "Terminado", ""),
    (71, "Dashboard y Analítica", "Gráficos e indicadores visuales",
     "Gráficas con Chart.js / Recharts para tendencias y comparativos",
     "Terminado", ""),
    (72, "Dashboard y Analítica", "Estadísticas mensuales",
     "Vista de estadísticas mensuales por módulo",
     "Terminado", ""),

    # EXPORTACIÓN / DOCUMENTOS
    (73, "Exportación y Documentos", "Generación de PDF",
     "PDF de facturas, contratos y órdenes con barcodes",
     "Terminado", ""),
    (74, "Exportación y Documentos", "Firma digital en documentos",
     "Integración de firma electrónica en contratos e facturas",
     "Terminado", ""),
    (75, "Exportación y Documentos", "Exportación de datos a Excel/CSV",
     "Descarga de listados de clientes, facturas e inventario",
     "Terminado", ""),

    # SEGURIDAD
    (76, "Seguridad", "Cifrado de contraseñas (bcrypt)",
     "Hash seguro de passwords antes de almacenar",
     "Terminado", ""),
    (77, "Seguridad", "Rate limiting",
     "Limitación de solicitudes para prevenir fuerza bruta",
     "Terminado", ""),
    (78, "Seguridad", "Cabeceras de seguridad (Helmet)",
     "Headers HTTP de seguridad configurados",
     "Terminado", ""),
    (79, "Seguridad", "Protección contra inyección SQL",
     "Consultas parametrizadas en toda la capa de datos",
     "Terminado", ""),
    (80, "Seguridad", "CORS configurado",
     "Control de orígenes permitidos para el API",
     "Terminado", ""),
]

# ─────────────────────────────────────────────
# ESTILOS
# ─────────────────────────────────────────────
COLOR_HEADER_BG    = "1F3864"   # azul oscuro
COLOR_HEADER_FONT  = "FFFFFF"   # blanco
COLOR_TITLE_BG     = "2E75B6"   # azul medio
COLOR_TERMINADO    = "E2EFDA"   # verde claro
COLOR_EN_PROCESO   = "FFF2CC"   # amarillo claro
COLOR_MOD_ODD      = "D6E4F0"   # azul muy claro (fila impar de módulo)
COLOR_MOD_EVEN     = "EBF5FB"   # azul muy muy claro (fila par de módulo)
COLOR_BORDER       = "B8CCE4"

thin = Side(style="thin", color=COLOR_BORDER)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def header_fill(color): return PatternFill("solid", fgColor=color)
def cell_fill(color):   return PatternFill("solid", fgColor=color)

def style_header(cell, font_size=11, bold=True, color=COLOR_HEADER_FONT, bg=COLOR_HEADER_BG):
    cell.font      = Font(name="Calibri", bold=bold, size=font_size, color=color)
    cell.fill      = header_fill(bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = BORDER

def style_data(cell, bg=None, bold=False, center=False):
    cell.font      = Font(name="Calibri", size=10, bold=bold)
    cell.alignment = Alignment(horizontal="center" if center else "left",
                               vertical="center", wrap_text=True)
    cell.border    = BORDER
    if bg:
        cell.fill = cell_fill(bg)

# ─────────────────────────────────────────────
# CREAR LIBRO
# ─────────────────────────────────────────────
wb = openpyxl.Workbook()

# ══════════════════════════════════════════════
# HOJA 1 — RESUMEN POR MÓDULO
# ══════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Resumen por Módulo"

# Título
ws1.merge_cells("A1:G1")
title_cell = ws1["A1"]
title_cell.value = f"INFORME DE FUNCIONALIDADES — {SISTEMA}"
title_cell.font      = Font(name="Calibri", bold=True, size=14, color=COLOR_HEADER_FONT)
title_cell.fill      = header_fill(COLOR_TITLE_BG)
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 30

ws1.merge_cells("A2:G2")
sub = ws1["A2"]
sub.value = f"Fecha del informe: {FECHA}     |     Total funcionalidades: {len(funcionalidades)}"
sub.font      = Font(name="Calibri", size=10, italic=True, color="444444")
sub.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[2].height = 18

# Encabezados columnas
headers = ["N°", "Módulo", "Funcionalidad", "Descripción", "Estado", "Observaciones"]
col_widths = [5, 28, 32, 55, 14, 40]

for col_i, (h, w) in enumerate(zip(headers, col_widths), start=1):
    cell = ws1.cell(row=3, column=col_i, value=h)
    style_header(cell)
    ws1.column_dimensions[get_column_letter(col_i)].width = w

ws1.row_dimensions[3].height = 22

# Datos
for row_i, (num, mod, func, desc, estado, obs) in enumerate(funcionalidades, start=4):
    bg_estado = COLOR_TERMINADO if estado == "Terminado" else COLOR_EN_PROCESO

    values = [num, mod, func, desc, estado, obs]
    for col_i, val in enumerate(values, start=1):
        cell = ws1.cell(row=row_i, column=col_i, value=val)
        center_cols = {1, 5}
        if col_i == 5:
            style_data(cell, bg=bg_estado, bold=True, center=True)
        elif col_i == 1:
            style_data(cell, center=True)
        else:
            style_data(cell)

    ws1.row_dimensions[row_i].height = 30

# Leyenda
leyenda_row = len(funcionalidades) + 5
ws1.merge_cells(f"A{leyenda_row}:B{leyenda_row}")
ws1[f"A{leyenda_row}"].value = "LEYENDA:"
ws1[f"A{leyenda_row}"].font  = Font(name="Calibri", bold=True, size=10)

cell_t = ws1.cell(row=leyenda_row, column=3, value="Terminado")
cell_t.fill = cell_fill(COLOR_TERMINADO)
cell_t.font = Font(name="Calibri", bold=True, size=10)
cell_t.border = BORDER

cell_p = ws1.cell(row=leyenda_row, column=4, value="En proceso")
cell_p.fill = cell_fill(COLOR_EN_PROCESO)
cell_p.font = Font(name="Calibri", bold=True, size=10)
cell_p.border = BORDER

ws1.freeze_panes = "A4"

# ══════════════════════════════════════════════
# HOJA 2 — TOTALES POR MÓDULO
# ══════════════════════════════════════════════
ws2 = wb.create_sheet("Totales por Módulo")

# Agregar datos por módulo
from collections import defaultdict
modulos = defaultdict(lambda: {"Terminado": 0, "En proceso": 0})
for _, mod, _, _, estado, _ in funcionalidades:
    modulos[mod][estado] += 1

ws2.merge_cells("A1:E1")
t2 = ws2["A1"]
t2.value = "TOTALES POR MÓDULO"
t2.font      = Font(name="Calibri", bold=True, size=13, color=COLOR_HEADER_FONT)
t2.fill      = header_fill(COLOR_TITLE_BG)
t2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 28

cols2   = ["Módulo", "Terminadas", "En Proceso", "Total", "% Completado"]
widths2 = [32, 14, 14, 10, 16]
for ci, (h, w) in enumerate(zip(cols2, widths2), 1):
    cell = ws2.cell(row=2, column=ci, value=h)
    style_header(cell)
    ws2.column_dimensions[get_column_letter(ci)].width = w
ws2.row_dimensions[2].height = 22

total_terminadas = total_proceso = 0
for ri, (mod, cnts) in enumerate(sorted(modulos.items()), start=3):
    t = cnts["Terminado"]
    p = cnts["En proceso"]
    tot = t + p
    pct = f"{round(t/tot*100)}%" if tot else "0%"
    total_terminadas += t
    total_proceso    += p

    bg = COLOR_MOD_ODD if ri % 2 == 0 else COLOR_MOD_EVEN
    for ci, val in enumerate([mod, t, p, tot, pct], 1):
        cell = ws2.cell(row=ri, column=ci, value=val)
        style_data(cell, bg=bg, center=(ci > 1))
    ws2.row_dimensions[ri].height = 22

# Fila total
total_row = len(modulos) + 3
for ci, val in enumerate(["TOTAL", total_terminadas, total_proceso,
                           total_terminadas + total_proceso,
                           f"{round(total_terminadas/(total_terminadas+total_proceso)*100)}%"], 1):
    cell = ws2.cell(row=total_row, column=ci, value=val)
    style_header(cell, font_size=10, bg="1F3864")

ws2.freeze_panes = "A3"

# ══════════════════════════════════════════════
# HOJA 3 — SÓLO EN PROCESO
# ══════════════════════════════════════════════
ws3 = wb.create_sheet("En Proceso")

ws3.merge_cells("A1:F1")
t3 = ws3["A1"]
t3.value = "FUNCIONALIDADES EN PROCESO"
t3.font      = Font(name="Calibri", bold=True, size=13, color="7F4000")
t3.fill      = header_fill("FFE699")
t3.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 28

col3_headers = ["N° Global", "Módulo", "Funcionalidad", "Descripción", "Estado", "Observaciones"]
col3_widths  = [10, 30, 35, 55, 14, 45]
for ci, (h, w) in enumerate(zip(col3_headers, col3_widths), 1):
    cell = ws3.cell(row=2, column=ci, value=h)
    style_header(cell, bg="BF8F00", font_size=10)
    ws3.column_dimensions[get_column_letter(ci)].width = w
ws3.row_dimensions[2].height = 22

en_proceso = [(n, m, f, d, e, o) for n, m, f, d, e, o in funcionalidades if e == "En proceso"]
if en_proceso:
    for ri, row in enumerate(en_proceso, start=3):
        for ci, val in enumerate(row, 1):
            cell = ws3.cell(row=ri, column=ci, value=val)
            style_data(cell, bg=COLOR_EN_PROCESO, center=(ci in {1, 5}), bold=(ci == 5))
        ws3.row_dimensions[ri].height = 30
else:
    ws3.merge_cells("A3:F3")
    ws3["A3"].value = "No hay funcionalidades en proceso actualmente."
    ws3["A3"].font  = Font(name="Calibri", italic=True, size=11)
    ws3["A3"].alignment = Alignment(horizontal="center", vertical="center")

ws3.freeze_panes = "A3"

# ─────────────────────────────────────────────
# GUARDAR
# ─────────────────────────────────────────────
output_path = "/home/user/ERP-PSI/Informe_Funcionalidades_ERP-PSI.xlsx"
wb.save(output_path)
print(f"✅ Informe generado: {output_path}")
print(f"   Total funcionalidades : {len(funcionalidades)}")
print(f"   Terminadas            : {sum(1 for *_, e, _ in funcionalidades if e == 'Terminado')}")
print(f"   En proceso            : {sum(1 for *_, e, _ in funcionalidades if e == 'En proceso')}")
